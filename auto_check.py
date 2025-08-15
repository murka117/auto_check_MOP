import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
import tkinter as tk
from tkinter import ttk, messagebox
import tkinter.filedialog as fd
from logic import clean_and_aggregate, build_final_table_multi

DATA_DIR = 'c:/auto_check_MOP/data'
RESULT_DIR = 'c:/auto_check_MOP/result'



# --- GUI класс ---
class App(tk.Tk):
    def apply_multiplier(self):
        # Сбросить все множители в 1
        for v in self.multipliers.values():
            v.set(1)
        # Применить множитель только к выбранному этажу
        floor = self.typical_floor.get()
        try:
            mult = int(self.typical_mult.get())
        except Exception:
            mult = 1
        if floor in self.multipliers:
            self.multipliers[floor].set(mult)
        self.recalc()
    def __init__(self):
        super().__init__()
        self.title('Проверка Excel по этажам')
        self.geometry('1200x700')
        self.configure(bg='#222')
        self.floors = {}
        self.multipliers = {}
        self.final_df = None
        self.typical_floor = tk.StringVar()
        self.typical_mult = tk.StringVar(value='1')
        # Кнопки выбора файлов
        btn_frame = tk.Frame(self, bg='#222')
        btn_frame.pack(pady=8)
        tk.Button(btn_frame, text='Открыть Excel', command=self.open_file, bg='#333', fg='#fff').pack(side='left', padx=5)
        # Множители
        mult_frame = tk.Frame(self, bg='#222')
        mult_frame.pack(pady=5)
        tk.Label(mult_frame, text='Типовой этаж:', fg='#fff', bg='#222').grid(row=0, column=0)
        floor_nums = []
        self.typical_floor_cb = ttk.Combobox(mult_frame, values=floor_nums, textvariable=self.typical_floor, state='readonly', width=8)
        self.typical_floor_cb.grid(row=0, column=1, padx=5)
        tk.Label(mult_frame, text='Множитель:', fg='#fff', bg='#222').grid(row=0, column=2)
        entry = tk.Entry(mult_frame, textvariable=self.typical_mult, width=5, bg='#333', fg='#fff', insertbackground='#fff')
        entry.grid(row=0, column=3, padx=5)
        tk.Button(mult_frame, text='Умножить', command=self.apply_multiplier, bg='#333', fg='#fff').grid(row=0, column=4, padx=5)
        # Кнопка экспорта в Excel
        export_btn = tk.Button(self, text='Экспортировать в Excel', command=self.export_to_excel, bg='#e8ffe8', fg='#222', relief='groove')
        export_btn.pack(pady=5)
        # Новый фрейм для таблицы
        self.tree_frame = tk.Frame(self)
        self.tree_frame.pack(fill='both', expand=True)
        # Кнопка для вывода промежуточных таблиц по этажам
        floor_btns_frame = tk.Frame(self)
        floor_btns_frame.pack(pady=2)
        # (Кнопки этажей появятся после загрузки файла)
    def open_file(self):
        path = fd.askopenfilename(title='Выберите Excel-файл', filetypes=[('Excel files', '*.xlsx *.xls')])
        if not path:
            return
        try:
            xl = pd.ExcelFile(path)
            self.floors = clean_and_aggregate(xl)
            # Множители по умолчанию = 1
            self.multipliers = {f: tk.IntVar(value=1) for f in self.floors}
            # Обновить список этажей в комбобоксе
            floor_nums = sorted([str(f) for f in self.floors if f not in ('0', '00')], key=lambda x: (len(x), x))
            self.typical_floor_cb['values'] = floor_nums
            if floor_nums:
                self.typical_floor.set(floor_nums[0])
            # Пересоздать кнопки этажей
            for widget in self.children.values():
                if isinstance(widget, tk.Frame):
                    for w in widget.winfo_children():
                        if isinstance(w, tk.Button) and w['text'].startswith('Показать этаж'):
                            w.destroy()
            btns_frame = None
            for widget in self.children.values():
                if isinstance(widget, tk.Frame):
                    for w in widget.winfo_children():
                        if isinstance(w, tk.Frame):
                            btns_frame = w
            if btns_frame:
                for f in floor_nums:
                    btn = tk.Button(btns_frame, text=f'Показать этаж {f}', command=lambda fl=f: self.show_floor_table(fl), width=14)
                    btn.pack(side='left', padx=2)
            # Пересчитать и показать таблицу
            self.recalc()
        except Exception as e:
            messagebox.showerror('Ошибка', f'Не удалось загрузить файл:\n{e}')

    def recalc(self):
        # Собрать множители для всех этажей (только числа)
        mults = {f: v.get() if hasattr(v, 'get') else v for f, v in self.multipliers.items()}
        self.final_df = build_final_table_multi(self.floors, mults)
        self.show_table()

    def show_table(self):
        # Удалить старый treeview
        for widget in self.tree_frame.winfo_children():
            widget.destroy()
        if self.final_df is None or self.final_df.empty:
            tk.Label(self.tree_frame, text='Нет данных для отображения.', fg='#222', bg='#fff').pack()
            return
        columns = list(self.final_df.columns)
        columns_with_idx = ['№'] + columns
        self.tree = ttk.Treeview(self.tree_frame, columns=columns_with_idx, show='headings', height=25)
        # Минималистичный Excel-like стиль
        style = ttk.Style(self)
        style.configure("Treeview.Heading", font=("Calibri", 11, "bold"), background="#f5f5f5", foreground="#222")
        style.configure("Treeview", font=("Calibri", 11), rowheight=22, borderwidth=0, background="#fff", fieldbackground="#fff")
        style.map("Treeview", background=[('selected', '#d0eaff')])
        # Заголовки
        for col in columns_with_idx:
            if col == '№':
                self.tree.heading(col, text=col, anchor='center')
                self.tree.column(col, width=40, anchor='center', minwidth=30, stretch=False)
            else:
                self.tree.heading(col, text=col, anchor='center')
                self.tree.column(col, width=110, anchor='center', minwidth=60, stretch=True)
        # Скроллы
        yscroll = ttk.Scrollbar(self.tree_frame, orient='vertical', command=self.tree.yview)
        xscroll = ttk.Scrollbar(self.tree_frame, orient='horizontal', command=self.tree.xview)
        self.tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        self.tree.pack(side='left', fill='both', expand=True)
        yscroll.pack(side='right', fill='y')
        xscroll.pack(side='bottom', fill='x')
        # Вставка данных: ошибки — розовый, успех — зелёный, остальное — белый
        for i, row in self.final_df.iterrows():
            vals = [i+1]
            for v in row:
                if isinstance(v, float):
                    vals.append(f'{v:.1f}')
                else:
                    vals.append(v)
            check = abs(row.get('Проверка', 0))
            if check > 1e-2:
                tag = 'err'
            elif check > 1e-6:
                tag = 'warn'
            else:
                tag = 'ok'
            self.tree.insert('', 'end', values=vals, tags=(tag,))
        self.tree.tag_configure('err', background='#ffe5ec')  # мягкий розовый
        self.tree.tag_configure('warn', background='#fffbe5') # мягкое желтое
        self.tree.tag_configure('ok', background='#e8ffe8')   # светло-зелёный
        # Выравнивание чисел вправо
        for col in columns_with_idx:
            if col not in ('Марка', 'Наименование', '№'):
                self.tree.column(col, anchor='e')

    def export_to_excel(self):
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill
        # Диалог выбора выходного файла
        self.output_path = fd.asksaveasfilename(title='Сохранить результат как...', defaultextension='.xlsx', filetypes=[('Excel files', '*.xlsx')])
        if not self.output_path:
            return
        wb = Workbook()
        ws = wb.active
        ws.title = 'Сводная'
        ws.append(list(self.final_df.columns))
        for i, row in self.final_df.iterrows():
            ws.append([v if not isinstance(v, float) else round(v, 1) for v in row])
        red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        for r in range(2, ws.max_row+1):
            val = ws.cell(row=r, column=ws.max_column).value
            try:
                if abs(float(val)) > 1e-6:
                    ws.cell(row=r, column=ws.max_column).fill = red_fill
            except:
                pass
        for f in sorted(self.floors, key=lambda x: (len(str(x)), str(x))):
            df = self.floors[f]
            ws_floor = wb.create_sheet(title=f'{f}')
            ws_floor.append(list(df.columns))
            for _, row in df.iterrows():
                ws_floor.append([v if not isinstance(v, float) else round(v, 1) for v in row])
        wb.save(self.output_path)
        messagebox.showinfo('Готово', f'Результат сохранён: {self.output_path}')

    def open_file(self):
        path = fd.askopenfilename(title='Выберите Excel-файл', filetypes=[('Excel files', '*.xlsx *.xls')])
        if not path:
            return
        try:
            xl = pd.ExcelFile(path)
            self.floors = clean_and_aggregate(xl)
            # Множители по умолчанию = 1
            self.multipliers = {f: tk.IntVar(value=1) for f in self.floors}
            # Обновить список этажей в комбобоксе
            floor_nums = sorted([str(f) for f in self.floors if f not in ('0', '00')], key=lambda x: (len(x), x))
            self.typical_floor_cb['values'] = floor_nums
            if floor_nums:
                self.typical_floor.set(floor_nums[0])
            # Пересоздать кнопки этажей
            for widget in self.children.values():
                if isinstance(widget, tk.Frame):
                    for w in widget.winfo_children():
                        if isinstance(w, tk.Button) and w['text'].startswith('Показать этаж'):
                            w.destroy()
            btns_frame = None
            for widget in self.children.values():
                if isinstance(widget, tk.Frame):
                    for w in widget.winfo_children():
                        if isinstance(w, tk.Frame):
                            btns_frame = w
            if btns_frame:
                for f in floor_nums:
                    btn = tk.Button(btns_frame, text=f'Показать этаж {f}', command=lambda fl=f: self.show_floor_table(fl), width=14)
                    btn.pack(side='left', padx=2)
            # Пересчитать и показать таблицу
            self.recalc()
        except Exception as e:
            messagebox.showerror('Ошибка', f'Не удалось загрузить файл:\n{e}')

# Запуск приложения
if __name__ == '__main__':
    app = App()
    app.mainloop()
